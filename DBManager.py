import openpyxl

class DBManager:
    def __init__(self, key):
        key = str(key)
        self.path = "C:\\Users\\abdel\\Downloads\\ews_barcode_scanner\\app\\server_files\\main_DB.xlsx"  # TODO: modify the path to be the actual path to the database sheet on the server (raspberry pi)
        self.wb_obj = openpyxl.load_workbook(self.path)
        self.sheet_obj = self.wb_obj.active
        rows_num = 0
        for i in range(2, self.sheet_obj.max_row):
            val = str(self.sheet_obj.cell(row=i, column=2).value)
            if val is None or len(val) <= 5:
                rows_num = i
                break
        if rows_num == 0:
            raise "Database unreadable."
        print(self.getName(key, rows_num)[0])
        self.updateDB(self.getName(key, rows_num)[1], 8)  # TODO: get the required quantity from the main manager class

    def getName(self, key, rows_num):
        for i in range(2, rows_num):
            if self.sheet_obj.cell(row=i, column=1).value == key:
                return [self.sheet_obj.cell(row=i, column=2).value, i]  # TODO: modify the returned value
        return "This piece of equipment is currently unavailable."
    
    def updateDB(self, row, required_quantity):  # TODO: pass in the required_quantity argument by the MainManager class
        print(int(self.sheet_obj.cell(row=row, column=6).value))
        remaining_quantity = int(self.sheet_obj.cell(row=row, column=6).value) - required_quantity
        print(remaining_quantity)  # debugging...
        self.sheet_obj.cell(row=row, column=6).value = str(remaining_quantity)
        self.wb_obj.save(self.path)
        self.wb_obj.close()


if __name__ == '__main__':
    DBManager("FRZ008")  # TODO: for testing purposes, remember to delete!